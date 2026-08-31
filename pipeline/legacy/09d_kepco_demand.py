# -*- coding: utf-8 -*-
"""09d_kepco_demand.py — 한전 시군 전력수요 산출 (19개 코드 체계)
=================================================================
원본: pipeline_out/kepco_usage/*.xlsx (한전 빅데이터센터 "산업분류별 전력사용량",
      2023.01~2025.12 3개년 누적, 로그인 다운로드 — 2026-07-06 확보분)
출력: pipeline_out/kepco_demand.json {sgg: {total_gwh_year, manuf_gwh_year, scope, files}}
방식: 3개년 누적 kWh ÷ 3 ÷ 1e6 = 연평균 GWh (구 사이트 build_kepco_usage.py와 동일 산식)
매핑: 19개 코드 — 시 단위 파일만 있는 천안·용인·안산은 구성 구 카드에 시 전체 값 배정
      (scope 각주 필수). 구 사이트의 '안산→시흥 합산' 방식은 폐기.
"""
import os, sys, io, glob, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import openpyxl

BASE = r"C:\Users\user\새 폴더"
UDIR = os.path.join(BASE, "pipeline_out", "kepco_usage")
OUT = os.path.join(BASE, "pipeline_out", "kepco_demand.json")

# 시군명 → (시 단위 키)
CITY = {"당진시": "당진", "서산시": "서산", "아산시": "아산", "천안시": "천안",
        "예산군": "예산", "홍성군": "홍성", "보령시": "보령", "화성시": "화성",
        "평택시": "평택", "용인시": "용인", "시흥시": "시흥", "안산시": "안산",
        "이천시": "이천", "파주시": "파주", "김포시": "김포"}
# 시 단위 키 → 배정 코드 + scope
ASSIGN = {
    "당진": (["44270"], "당진시"), "서산": (["44210"], "서산시"),
    "아산": (["44200"], "아산시"), "보령": (["44180"], "보령시"),
    "홍성": (["44800"], "홍성군"), "예산": (["44810"], "예산군"),
    "화성": (["41590"], "화성시"), "평택": (["41220"], "평택시"),
    "시흥": (["41390"], "시흥시"), "이천": (["41500"], "이천시"),
    "파주": (["41480"], "파주시"), "김포": (["41570"], "김포시"),
    "천안": (["44131", "44133"], "천안시 전체(동남+서북 합산)"),
    "용인": (["41463", "41461", "41465"], "용인시 전체(3개 구 합산)"),
    "안산": (["41271", "41273"], "안산시 전체(상록+단원 합산)"),
}


# 부분 export 의심 — 카드 배정 보류 (2026-07-14 상식 점검: 호수 783k가 안산(1)·용인(1) 등
# '구 1개 분량'과 동급, 인구 65만 도시가 아산의 1/5 사용량. 구별 (1)/(2) 재다운로드 필요)
SUSPECT = {"천안": "부분 export 의심 — 천안시 (1)/(2) 구별 파일 재다운로드 필요"}


def to_num(c):
    return float(str(c).replace(",", "").strip()) if c not in (None, "") else 0.0


def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    manuf = sum(to_num(r[4]) for r in rows[4:] if (r[2] or "").strip() == "제조업")
    total = to_num(rows[-1][4])
    return total, manuf


city_kwh = {}
files = {}
for p in sorted(glob.glob(os.path.join(UDIR, "*.xlsx"))):
    city = CITY.get(os.path.basename(p).split(" ")[0])
    if not city:
        print("미매핑:", os.path.basename(p))
        continue
    t, m = parse(p)
    a, b = city_kwh.get(city, (0.0, 0.0))
    city_kwh[city] = (a + t, b + m)
    files.setdefault(city, []).append(os.path.basename(p))

out = {"period_label": "2023.01–2025.12(3개년 누적치의 연평균)",
       "source_label": "한국전력공사 빅데이터센터(bigdata.kepco.co.kr), 산업분류별 전력사용량 2023–2025 연평균",
       "caveat": "시군구 전체 업종(KSIC) 합산치로, 산단 단위 추정 소비량과 1:1 비교 대상이 아님",
       "by_code": {}}
for city, (t, m) in city_kwh.items():
    if city in SUSPECT:
        print(f"{city}: 배정 보류 — {SUSPECT[city]} (산출값 {t/3/1e6:,.0f} GWh 참고용)")
        continue
    codes, scope = ASSIGN[city]
    for code in codes:
        out["by_code"][code] = {
            "total_gwh_year": round(t / 3 / 1e6, 2),
            "manuf_gwh_year": round(m / 3 / 1e6, 2),
            "scope": scope, "files": files[city]}
    print(f"{city}({scope}): {t/3/1e6:,.0f} GWh/년 → {codes}")

json.dump(out, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"\nkepco_demand.json — {len(out['by_code'])}개 코드")
